from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def copy_cell_value(source_file, source_sheet, source_cell, destination_file, destination_sheet, destination_cell):
    # Загрузка исходного файла Excel
    source_wb = load_workbook(source_file)
    source_ws = source_wb[source_sheet]

    # Загрузка файла Excel назначения
    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb[destination_sheet]

    # Получение значения из исходной ячейки
    source_value = source_ws[source_cell].value

    # Помещение значения в ячейку назначения
    dest_ws[destination_cell] = source_value

    # Сохранение изменений в файле Excel назначения
    dest_wb.save(destination_file)

    # Закрытие файлов
    source_wb.close()
    dest_wb.close()

# Сопоставление между исходными файлами Excel и их целевыми ячейками
mapping = {
    ("1_PN.xlsx", "свод", "C10"): ("ZAGAL_DATA.xlsx", "Sheet1", "D5"),
    ("1_PN.xlsx", "свод", "C12"): ("ZAGAL_DATA.xlsx", "Sheet1", "G5"),
    ("1_PN.xlsx", "свод", "C15"): ("ZAGAL_DATA.xlsx", "Sheet1", "R5"),
    ("1_PN.xlsx", "свод", "C19"): ("ZAGAL_DATA.xlsx", "Sheet1", "S5"),
    ("1_PN.xlsx", "свод", "C24"): ("ZAGAL_DATA.xlsx", "Sheet1", "O5"),
    ("U_V.xlsx", "Шаблон", "G9"): ("ZAGAL_DATA.xlsx", "Sheet1", "Q5"),
    ("U_V.xlsx", "Шаблон", "AA9"): ("ZAGAL_DATA.xlsx", "Sheet1", "T5"),
    ("O_S_K_R.xlsx", "Лист1", "E10"): ("ZAGAL_DATA.xlsx", "Sheet1", "P5"),
    ("K_B_P_S.xlsx", "Лист1", "B7"): ("ZAGAL_DATA.xlsx", "Sheet1", "E5"),
    ("N_P_O_N_O.xlsx", "Шаблон", "D9"): ("ZAGAL_DATA.xlsx", "Sheet1", "F5"),
    ("N_P_O_N_O_35.xlsx", "Шаблон", "D9"): ("ZAGAL_DATA.xlsx", "Sheet1", "L5"),
    ("N_P_O_N_O_inval.xlsx", "Шаблон", "D9"): ("ZAGAL_DATA.xlsx", "Sheet1", "J5"),
    ("N_P_Z_B_za_SZ_SB.xlsx", "Шаблон", "J8"): ("ZAGAL_DATA.xlsx", "Sheet1", "N5"),
    ("N_P_Z_B_35_za_SZ_SB.xlsx", "Шаблон", "F8"): ("ZAGAL_DATA.xlsx", "Sheet1", "M5"),
    ("N_P_Z_B_inval_za_SZ_SB.xlsx", "Шаблон", "F8"): ("ZAGAL_DATA.xlsx", "Sheet1", "K5"),
}

# Копирование значений по заданному сопоставлению
for source, destination in mapping.items():
    source_file, source_sheet, source_cell = source
    destination_file, destination_sheet, destination_cell = destination
    copy_cell_value(source_file, source_sheet, source_cell, destination_file, destination_sheet, destination_cell)

print("Значения скопированы успешно!")
